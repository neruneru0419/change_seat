from flask import Flask
from flask import render_template, request, send_file
from flask_cors import CORS

import openpyxl as excel
import random 
import datetime

XLSX_MIMETYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
app = Flask(__name__, static_folder="./build/static", template_folder="./build")
CORS(app)

row = 7
column = 8
def change_sheet(class_member, memberSize):
    class_number_member = []
    lst = class_member.split(",")
    i = 0
    print(len(lst))
    while i < len(lst): 
        class_number_member.append(lst[i] + lst[i+1])
        i += 2
    random.shuffle(class_number_member)
    for i in range(memberSize - len(class_number_member)):
        class_number_member.append("")
    
    print("メンバー"+ str(class_number_member))
    return str(class_number_member)

def reverse_sheet(class_member):
    lst = class_member.split(",")
    lst.reverse()
    print("メンバー"+ str(class_number_member))
    return str(lst)

def make_excel(class_number_member):
    wb = excel.Workbook()
    ws = wb.active
    cnt = 0
    ws.cell(column=(column/2), row=1, value="黒板")
    for i in range(2, ((row + 2) * 2), 2):
        for j in range(1, column):
            class_number_list = class_number_member[cnt].strip()[0:6]
            member_list = class_number_member[cnt].strip()[6:]
            ws.cell(column=j, row=i, value=class_number_list)
            ws.cell(column=j, row=i+1, value=member_list)
            cnt += 1
    wb.save("class.xlsx")

@app.route('/', methods=["GET", "POST"])
def index():
    if request.method == "GET":
            return change_sheet(request.args.get('member'), int(request.args.get('memberSize')))

@app.route('/reverse')
def reverse():
    return reverse_sheet(request.args.get('member'))

@app.route('/download')
def excel_dl():
    now_datetime = datetime.datetime.now()
    load_file_name = 'class.xlsx'
    download_file_name = str(now_datetime.year) + "/" + str(now_datetime.month) + "/" \
                        + str(now_datetime.day) + "_" + str(now_datetime.hour) + ":" \
                        + str(now_datetime.minute) + ":" + str(now_datetime.second) + ".xlsx"""
    make_excel()
    return send_file(load_file_name, as_attachment = True, \
        attachment_filename = download_file_name, \
        mimetype = XLSX_MIMETYPE)

if __name__ == '__main__':
    app.run(host="127.0.0.1", port=8080, debug=True)
    #make_excel(7, 8)

